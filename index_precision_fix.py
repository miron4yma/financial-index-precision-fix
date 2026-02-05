# -*- coding: utf-8 -*-
"""
Financial Index Precision Fix (High-Precision Adjustment Calculator)
====================================================================
Author: Yuri Miron
Description:
    Solves floating-point truncation issues (IEEE 754) in legacy financial systems.
    Calculates the exact percentage adjustment 'p' required to transform a 
    Base Quantity (B) into a Theoretical Quantity (T) after truncation.
    
    Formula: Quantity_Final = TRUNC( Base * (1 + p) )
    
    This script ensures that 'p' is calculated with arbitrary precision (Decimal)
    and includes an 'epsilon' safe-guard to prevent off-by-one errors during 
    system truncation.

Features:
    - Uses 'decimal' library for arbitrary precision (50 places).
    - Implements a 'truncation-safe' logic to find the smallest valid 'p'.
    - Auto-detects headers in input files (Excel/CSV resilience).
    - Generates a detailed audit log (Excel) with proof of calculation.
"""

from __future__ import annotations
from decimal import Decimal, getcontext, ROUND_UP
import pandas as pd
import xlrd
import re, datetime, argparse
from pathlib import Path
from typing import Optional, Tuple

# Set precision to 50 decimal places to avoid standard float errors
getcontext().prec = 50

# ---- CLI Arguments ----
parser = argparse.ArgumentParser(description='Financial Index Precision Calculator')
parser.add_argument('--base', default='assets_base.xlsx', help="Base Quantity File (Current Portfolio)")
parser.add_argument('--secondary', default='assets_secondary.xlsx', help="Secondary Assets File (e.g. DRs)")
parser.add_argument('--target', default='target_portfolios.xlsx', help="Target Theoretical Portfolios")
parser.add_argument('--output', default='adjustment_results.xlsx', help='Output Excel File')
parser.add_argument('--no-cover', action='store_true', help='Skip instruction cover sheet')
args = parser.parse_args()

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_FILE = (SCRIPT_DIR / args.base).resolve()
SEC_FILE = (SCRIPT_DIR / args.secondary).resolve()
TARGET_FILE = (SCRIPT_DIR / args.target).resolve()
OUTPUT_FILE = (SCRIPT_DIR / args.output).resolve()

DECIMAL_PLACES = 13  # Standard for financial transaction systems

# ---- Utils ----
now = lambda: datetime.datetime.now().strftime('%H:%M:%S')
log = lambda m: print(f"[{now()}] {m}")
norm = lambda s: re.sub(r'\s+', ' ', str(s).strip()).lower()
from unicodedata import normalize as _ud_norm
noacc = lambda s: _ud_norm('NFKD', str(s)).encode('ASCII','ignore').decode('ASCII').lower()

def safe_decimal_int(x) -> Optional[int]:
    """Safely converts a value to int using Decimal to avoid float artifacts."""
    try:
        if pd.isna(x):
            return None
        return int(Decimal(str(x)))
    except Exception:
        return None

def calculate_safe_adjustment(B: int, T: int, N: int = DECIMAL_PLACES) -> Tuple[Optional[Decimal], Optional[int], bool]:
    """
    Core Logic: Finds the smallest 'p' (percentage) that satisfies:
    TRUNC( B * (1 + p) ) == T
    
    It adds an 'epsilon' to the mathematical lower bound (T/B - 1) 
    to ensure the value survives system truncation.
    """
    try:
        if B is None or T is None:
            return None, None, False
        B = int(B); T = int(T)
        if B == 0:
            return None, None, False
            
        Bd = Decimal(B); Td = Decimal(T)
        
        # Define step (quantum) and safety epsilon
        q = Decimal('1e-' + str(N))
        EPS = Decimal('1e-' + str(N+2))
        
        # Theoretical lower bound
        lower_bound = (Td / Bd) - Decimal(1)
        
        # Candidate p = ceil(lower_bound + epsilon)
        candidate_p = (lower_bound + EPS).quantize(q, rounding=ROUND_UP)
        
        # Validation Trial
        trial_qty = Bd * (Decimal(1) + candidate_p)
        
        # If trial fails (due to internal precision limits), bump p by one quantum
        if int(trial_qty) < int(Td):
            candidate_p = (candidate_p + q).quantize(q, rounding=ROUND_UP)
            trial_qty = Bd * (Decimal(1) + candidate_p)
            
        return candidate_p, int(trial_qty), int(trial_qty) == T
    except Exception:
        return None, None, False

# ---- 1. Load Base Assets (Current Positions) ----
log(f"Loading Base Assets: {BASE_FILE.name}")
try:
    df_base = pd.read_excel(BASE_FILE, sheet_name=0, engine='openpyxl')
except FileNotFoundError:
    log("Base file not found. Please provide a valid .xlsx file.")
    exit(1)

# Dynamic Column Mapping (Agnostic to file format)
colmap = {norm(c): c for c in df_base.columns}
col_code = colmap.get('ticker') or colmap.get('symbol') or colmap.get('code')
# Logic to find the quantity column (looks for 'qty', 'shares', 'position')
col_qty = None
for k, v in colmap.items():
    if ('qty' in k) or ('quantity' in k) or ('position' in k):
        col_qty = v; break

if not col_code or not col_qty:
    # Fallback for demo purposes if columns are missing
    log("Warning: Specific columns not found, using generic column indices 0 and 1.")
    df_base.columns = ['Ticker', 'Qty'] + list(df_base.columns[2:])
    col_code = 'Ticker'
    col_qty = 'Qty'

df_base['NORMALIZED_CODE'] = df_base[col_code].astype(str).str.strip()
df_base['BASE_QTY'] = df_base[col_qty].apply(safe_decimal_int)
base_lookup = dict(zip(df_base['NORMALIZED_CODE'], df_base['BASE_QTY']))
log(f"Base Assets Loaded: {len(base_lookup)} records.")

# ---- 2. Load Secondary Assets (Optional - e.g. DRs/Receipts) ----
base_sec_lookup = {}
if SEC_FILE.exists():
    log(f"Loading Secondary Assets: {SEC_FILE.name}")
    df_sec = pd.read_excel(SEC_FILE, sheet_name=0, engine='openpyxl')
    # ... (Similar dynamic mapping logic would go here) ...
    # Simplified for the portfolio version:
    colmap_sec = {norm(c): c for c in df_sec.columns}
    col_code_sec = colmap_sec.get('ticker') or list(df_sec.columns)[0]
    col_qty_sec = colmap_sec.get('qty') or list(df_sec.columns)[-1]
    
    df_sec['NORMALIZED_CODE'] = df_sec[col_code_sec].astype(str).str.strip()
    df_sec['BASE_QTY'] = df_sec[col_qty_sec].apply(safe_decimal_int)
    base_sec_lookup = dict(zip(df_sec['NORMALIZED_CODE'], df_sec['BASE_QTY']))

# ---- 3. Process Target Portfolios ----
log(f"Processing Target Portfolios: {TARGET_FILE.name}")
if not TARGET_FILE.exists():
    log("Target file not found. Creating dummy output for demonstration.")
    sheet_names = []
else:
    xls = pd.ExcelFile(str(TARGET_FILE), engine='openpyxl')
    sheet_names = xls.sheet_names

writer = pd.ExcelWriter(str(OUTPUT_FILE), engine='openpyxl')
summary_stats = []

for sheet in sheet_names:
    log(f"Analyzing Portfolio: {sheet}")
    df_target = pd.read_excel(TARGET_FILE, sheet_name=sheet, engine='openpyxl')
    
    # Header Detection Logic (finds where the data actually starts)
    header_idx = 0
    for i in range(min(20, len(df_target))):
        row_str = str(df_target.iloc[i].values).lower()
        if 'ticker' in row_str or 'symbol' in row_str:
            header_idx = i; break
            
    df_target = pd.read_excel(TARGET_FILE, sheet_name=sheet, header=header_idx, engine='openpyxl')
    
    # Normalize Columns
    t_map = {norm(c): c for c in df_target.columns}
    t_code = t_map.get('ticker') or t_map.get('symbol')
    t_qty = t_map.get('quantity') or t_map.get('qty') or t_map.get('theoretical')
    
    if not t_code or not t_qty:
        continue

    results = []
    success_count = 0
    
    for _, row in df_target.iterrows():
        ticker = str(row.get(t_code, '')).strip()
        target_qty = safe_decimal_int(row.get(t_qty))
        
        if not ticker or target_qty is None:
            continue
            
        # Lookup Base Quantity (Primary or Secondary)
        base_qty = base_lookup.get(ticker)
        source = 'PRIMARY'
        if base_qty is None:
            base_qty = base_sec_lookup.get(ticker)
            source = 'SECONDARY'
            
        if base_qty is None:
            results.append({
                'Ticker': ticker, 'Base_Qty': 'N/A', 'Target_Qty': target_qty, 
                'Adjustment_Factor': 'N/A', 'Status': 'MISSING_BASE'
            })
            continue
            
        # Calculate!
        adj_factor, proof_qty, is_valid = calculate_safe_adjustment(base_qty, target_qty)
        
        if is_valid:
            success_count += 1
            
        results.append({
            'Ticker': ticker,
            'Base_Qty': base_qty,
            'Target_Qty': target_qty,
            'Adjustment_Factor': float(adj_factor) if adj_factor else None,
            'Proof_Check': proof_qty,
            'Status': 'OK' if is_valid else 'FAIL'
        })
        
    pd.DataFrame(results).to_excel(writer, sheet_name=sheet[:31], index=False)
    summary_stats.append({'Portfolio': sheet, 'Success': success_count, 'Total': len(results)})

# Save Summary
if summary_stats:
    pd.DataFrame(summary_stats).to_excel(writer, sheet_name='SUMMARY', index=False)

# ---- 4. Create Instruction Cover Sheet (Documentation) ----
if not args.no_cover:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Save first to allow loading
        writer.close() 
        wb = load_workbook(OUTPUT_FILE)
        
        if 'Instructions' in wb.sheetnames:
            wb.remove(wb['Instructions'])
        ws = wb.create_sheet('Instructions', 0)
        
        # Style
        header_font = Font(color='FFFFFF', bold=True, size=18)
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid') # Dark Blue
        
        ws['A1'] = "FINANCIAL INDEX PRECISION FIX - DOCUMENTATION"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E2')
        
        ws['B4'] = "Mathematical Logic:"
        ws['B4'].font = Font(bold=True, size=12)
        
        explanation = [
            "1. Problem: Legacy systems truncate decimal values, causing off-by-one errors.",
            "2. Goal: Find 'p' such that TRUNC( Base * (1 + p) ) == Target.",
            "3. Solution: Calculate p = (Target / Base) - 1 + Epsilon.",
            "4. Safety: 'Epsilon' ensures the float representation is slightly above the mathematical threshold."
        ]
        
        for i, line in enumerate(explanation):
            ws.cell(row=5+i, column=2).value = line
            
        wb.save(OUTPUT_FILE)
        log("Documentation cover sheet added.")
    except Exception as e:
        log(f"Cover sheet generation skipped: {e}")
        # If cover fails, just close the writer if not closed
        try: writer.close() 
        except: pass
else:
    writer.close()

log(f"Process Complete. Results saved to: {OUTPUT_FILE.name}")